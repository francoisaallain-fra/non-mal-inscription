import csv
import html
import json
import math
import unicodedata
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
INSEE_XLSX = ROOT / "data/01_raw/insee/donnees_insee_premiere_n1986.xlsx"
PRESIDENTIELLE_T1 = (
    ROOT
    / "data/01_raw/ministere_interieur/2022-presidentielle-1-bureau_de_vote.csv"
)
DEPARTEMENTS_GEOJSON = ROOT / "data/04_analysis/geodata/departements-2022.geojson"

MAL_INSCRIPTION_CSV = (
    ROOT / "data/04_analysis/insee/2022-mal-inscription-departements.csv"
)
MELENCHON_CSV = (
    ROOT
    / "data/04_analysis/elections/2022-presidentielle-melenchon-departements-t1.csv"
)
JOINED_GEOJSON = (
    ROOT
    / "data/04_analysis/cartographie/2022-departements-mal-inscription-melenchon.geojson"
)
QA_UNMATCHED = (
    ROOT
    / "data/04_analysis/qa/2022-departements-non-apparies-melenchon.csv"
)
QA_TOTALS = (
    ROOT
    / "data/04_analysis/qa/2022-controles-totaux-melenchon-departements.csv"
)

MAPS = {
    "mal": ROOT / "maps/2022-mal-inscrits-departements-presidentielle.html",
    "melenchon": ROOT
    / "maps/2022-vote-melenchon-presidentielle-departements.html",
    "cross": ROOT
    / "maps/2022-croisement-mal-inscrits-melenchon-presidentielle-departements.html",
    "cross_web": ROOT
    / "maps/2022-croisement-mal-inscrits-melenchon-presidentielle-departements-web.html",
}

DEPARTEMENTS_OUTREMER = {
    "ZA": "971",
    "ZB": "972",
    "ZC": "973",
    "ZD": "974",
    "ZM": "976",
    "ZN": "988",
    "ZP": "987",
    "ZS": "975",
    "ZW": "986",
}

TERRITORY_INSET_BOXES = {
    "975": (-5.7, 43.55, -4.9, 44.15),
    "986": (-5.7, 42.75, -4.9, 43.35),
    "987": (-4.75, 43.55, -3.95, 44.15),
    "988": (-4.75, 42.75, -3.95, 43.35),
    "ZX": (-3.8, 43.55, -3.0, 44.15),
    "ZZ": (-3.8, 42.75, -3.0, 43.35),
}

BASE_HEADERS = {
    "Code de la circonscription",
    "Code de la commune",
    "Code du b.vote",
    "Inscrits",
    "Votants",
    "Blancs",
    "Nuls",
    "Exprimés",
}
CANDIDATE_HEADERS = {
    "N°Panneau",
    "N.Pan.",
    "Sexe",
    "Nom",
    "Prénom",
    "Voix",
    "% Voix/Ins",
    "% Voix/Exp",
}


def ensure_dirs():
    for path in [
        MAL_INSCRIPTION_CSV,
        MELENCHON_CSV,
        JOINED_GEOJSON,
        QA_UNMATCHED,
        QA_TOTALS,
        *MAPS.values(),
    ]:
        path.parent.mkdir(parents=True, exist_ok=True)


def code_departement(value):
    value = str(value).strip()
    return DEPARTEMENTS_OUTREMER.get(value, value.zfill(2))


def to_int(value):
    value = str(value or "").strip().replace(" ", "")
    return int(value) if value else 0


def slug(value):
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    ascii_text = "".join(char for char in normalized if not unicodedata.combining(char))
    return ascii_text.upper().strip()


def extract_mal_inscription():
    if MAL_INSCRIPTION_CSV.exists():
        with MAL_INSCRIPTION_CSV.open("r", encoding="utf-8", newline="") as stream:
            rows = list(csv.DictReader(stream))
        return {row["code_departement"]: row for row in rows}

    from openpyxl import load_workbook

    wb = load_workbook(INSEE_XLSX, data_only=True)
    ws = wb["Figure 4"]

    rows = []
    for cells in ws.iter_rows(min_row=4, values_only=True):
        code, name, part_other_commune, part_same_commune = cells[:4]
        if not code or not name:
            continue
        rows.append(
            {
                "code_departement": code_departement(code),
                "libelle_departement": str(name).strip(),
                "part_mal_inscrits": f"{float(part_other_commune) / 100:.6f}",
                "part_inscrits_commune_residence": f"{float(part_same_commune) / 100:.6f}",
                "source": "INSEE Première n°1986, Figure 4",
            }
        )

    with MAL_INSCRIPTION_CSV.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)
    return {row["code_departement"]: row for row in rows}


def election_layout(headers):
    first_candidate_idx = min(
        index
        for index, header in enumerate(headers)
        if header in CANDIDATE_HEADERS and header not in BASE_HEADERS
    )
    candidate_headers = headers[first_candidate_idx:]
    return {
        "departement_idx": headers.index("Code du département"),
        "libelle_departement_idx": headers.index("Libellé du département"),
        "bureau_identity_indices": [
            headers.index("Code de la circonscription"),
            headers.index("Code de la commune"),
            headers.index("Code du b.vote"),
        ],
        "inscrits_idx": headers.index("Inscrits"),
        "votants_idx": headers.index("Votants"),
        "blancs_idx": headers.index("Blancs"),
        "nuls_idx": headers.index("Nuls"),
        "exprimes_idx": headers.index("Exprimés"),
        "first_candidate_idx": first_candidate_idx,
        "block_width": len(candidate_headers),
        "nom_offset": candidate_headers.index("Nom"),
        "prenom_offset": candidate_headers.index("Prénom"),
        "voix_offset": candidate_headers.index("Voix"),
    }


def candidate_blocks(line, layout):
    first = layout["first_candidate_idx"]
    width = layout["block_width"]
    for offset in range(first, len(line), width):
        if offset + width > len(line):
            continue
        yield {
            "nom": line[offset + layout["nom_offset"]].strip(),
            "prenom": line[offset + layout["prenom_offset"]].strip(),
            "voix": line[offset + layout["voix_offset"]].strip(),
        }


def is_melenchon(candidate):
    return slug(candidate["nom"]) == "MELENCHON" and slug(candidate["prenom"]) in {
        "JEAN-LUC",
        "JEAN LUC",
    }


def aggregate_melenchon_t1():
    if not PRESIDENTIELLE_T1.exists() and MELENCHON_CSV.exists():
        with MELENCHON_CSV.open("r", encoding="utf-8", newline="") as stream:
            rows = list(csv.DictReader(stream))
        return {row["code_departement"]: row for row in rows}
    if not PRESIDENTIELLE_T1.exists():
        raise FileNotFoundError(
            "CSV brut présidentielle 2022 T1 introuvable. "
            "Téléchargez-le depuis l'URL Hexagonal/data.gouv référencée pour "
            "2022-presidentielle-1-bureau_de_vote.csv, ou fournissez l'agrégat "
            f"{MELENCHON_CSV}."
        )

    totals = {}
    seen_bureaux_by_dep = {}
    with PRESIDENTIELLE_T1.open("r", encoding="latin1", newline="") as stream:
        reader = csv.reader(stream, delimiter=";")
        layout = election_layout(next(reader))
        for line in reader:
            dep = code_departement(line[layout["departement_idx"]])
            seen_bureaux = seen_bureaux_by_dep.setdefault(dep, set())
            bureau_key = tuple(
                line[index].strip() for index in layout["bureau_identity_indices"]
            )
            if bureau_key in seen_bureaux:
                continue
            seen_bureaux.add(bureau_key)

            bucket = totals.setdefault(
                dep,
                {
                    "code_departement": dep,
                    "libelle_departement": line[layout["libelle_departement_idx"]],
                    "methode_melenchon": "calcul_candidat_nom_prenom",
                    "source_melenchon": (
                        "Ministère de l'Intérieur via Hexagonal, présidentielle 2022 T1"
                    ),
                    "source_exprimes": "Ministère de l'Intérieur, présidentielle 2022 T1",
                    "bureaux": 0,
                    "inscrits": 0,
                    "votants": 0,
                    "blancs": 0,
                    "nuls": 0,
                    "exprimes": 0,
                    "voix_melenchon": 0,
                    "bureaux_avec_melenchon": 0,
                },
            )
            bucket["bureaux"] += 1
            for field in ["inscrits", "votants", "blancs", "nuls", "exprimes"]:
                bucket[field] += to_int(line[layout[f"{field}_idx"]])

            found = False
            for candidate in candidate_blocks(line, layout):
                if is_melenchon(candidate):
                    bucket["voix_melenchon"] += to_int(candidate["voix"])
                    found = True
                    break
            if found:
                bucket["bureaux_avec_melenchon"] += 1

    rows = []
    for _, row in sorted(totals.items()):
        rows.append(
            {
                **row,
                "part_melenchon_exprimes": (
                    f"{row['voix_melenchon'] / row['exprimes']:.6f}"
                    if row["exprimes"]
                    else ""
                ),
                "part_melenchon_inscrits": (
                    f"{row['voix_melenchon'] / row['inscrits']:.6f}"
                    if row["inscrits"]
                    else ""
                ),
            }
        )

    with MELENCHON_CSV.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)

    national = {
        "bureaux": sum(row["bureaux"] for row in rows),
        "inscrits": sum(row["inscrits"] for row in rows),
        "exprimes": sum(row["exprimes"] for row in rows),
        "voix_melenchon": sum(row["voix_melenchon"] for row in rows),
    }
    with QA_TOTALS.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(
            stream,
            fieldnames=[
                "scope",
                "bureaux",
                "inscrits",
                "exprimes",
                "voix_melenchon",
                "part_melenchon_exprimes",
                "methode",
            ],
        )
        writer.writeheader()
        writer.writerow(
            {
                "scope": "France entière source ministère T1",
                **national,
                "part_melenchon_exprimes": f"{national['voix_melenchon'] / national['exprimes']:.6f}",
                "methode": "agrégation par département du candidat Mélenchon",
            }
        )
    return {row["code_departement"]: row for row in rows}


def percentile_scores(features, key):
    values = sorted(
        (feature["properties"].get(key), index)
        for index, feature in enumerate(features)
        if isinstance(feature["properties"].get(key), (int, float))
    )
    scores = {}
    if len(values) == 1:
        return {values[0][1]: 1.0}
    for rank, (_, index) in enumerate(values):
        scores[index] = rank / (len(values) - 1)
    return scores


def enrich_geojson(mal_by_dep, melenchon_by_dep):
    geojson = json.load(DEPARTEMENTS_GEOJSON.open(encoding="utf-8"))
    unmatched = []
    geo_codes = {feature["properties"].get("code") for feature in geojson["features"]}

    for feature in geojson["features"]:
        props = feature["properties"]
        dep = props.get("code")
        props["code_departement"] = dep
        props["libelle_departement"] = props.get("nom")

        mal = mal_by_dep.get(dep)
        vote = melenchon_by_dep.get(dep)
        if mal:
            props["part_mal_inscrits"] = float(mal["part_mal_inscrits"])
        else:
            unmatched.append({"code_departement": dep, "dataset": "mal_inscription"})

        if vote:
            for key in [
                "bureaux",
                "inscrits",
                "exprimes",
                "voix_melenchon",
                "bureaux_avec_melenchon",
            ]:
                props[key] = int(vote[key])
            props["methode_melenchon"] = vote["methode_melenchon"]
            props["source_melenchon"] = vote["source_melenchon"]
            props["source_exprimes"] = vote["source_exprimes"]
            if vote["part_melenchon_exprimes"]:
                props["part_melenchon_exprimes"] = float(vote["part_melenchon_exprimes"])
            if vote["part_melenchon_inscrits"]:
                props["part_melenchon_inscrits"] = float(vote["part_melenchon_inscrits"])
        else:
            unmatched.append({"code_departement": dep, "dataset": "vote_melenchon"})

    for dep, vote in sorted(melenchon_by_dep.items()):
        if dep in geo_codes or dep not in TERRITORY_INSET_BOXES:
            continue
        x1, y1, x2, y2 = TERRITORY_INSET_BOXES[dep]
        props = {
            "code": dep,
            "nom": vote["libelle_departement"],
            "code_departement": dep,
            "libelle_departement": vote["libelle_departement"],
            "territoire_schematique": True,
            "absence_mal_inscription": "Non disponible dans l'étude INSEE n°1986",
        }
        for key in [
            "bureaux",
            "inscrits",
            "exprimes",
            "voix_melenchon",
            "bureaux_avec_melenchon",
        ]:
            props[key] = int(vote[key])
        props["methode_melenchon"] = vote["methode_melenchon"]
        props["source_melenchon"] = vote["source_melenchon"]
        props["source_exprimes"] = vote["source_exprimes"]
        if vote["part_melenchon_exprimes"]:
            props["part_melenchon_exprimes"] = float(vote["part_melenchon_exprimes"])
        if vote["part_melenchon_inscrits"]:
            props["part_melenchon_inscrits"] = float(vote["part_melenchon_inscrits"])
        geojson["features"].append(
            {
                "type": "Feature",
                "properties": props,
                "geometry": {
                    "type": "Polygon",
                    "coordinates": [
                        [[x1, y1], [x2, y1], [x2, y2], [x1, y2], [x1, y1]]
                    ],
                },
            }
        )
        unmatched.append({"code_departement": dep, "dataset": "mal_inscription"})

    mal_scores = percentile_scores(geojson["features"], "part_mal_inscrits")
    vote_scores = percentile_scores(geojson["features"], "part_melenchon_exprimes")
    for index, feature in enumerate(geojson["features"]):
        props = feature["properties"]
        mal_score = mal_scores.get(index)
        vote_score = vote_scores.get(index)
        if mal_score is None or vote_score is None:
            continue
        props["score_mal_inscription"] = round(mal_score, 6)
        props["score_melenchon"] = round(vote_score, 6)
        props["score_croise"] = round(math.sqrt(mal_score * vote_score), 6)
        props["methode_score_croise"] = "moyenne_geometrique_des_rangs_percentiles"

    with QA_UNMATCHED.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=["code_departement", "dataset"])
        writer.writeheader()
        writer.writerows(unmatched)
    return geojson


def hex_to_rgb(color):
    color = color.lstrip("#")
    return tuple(int(color[i : i + 2], 16) for i in (0, 2, 4))


def rgb_to_hex(rgb):
    return "#" + "".join(f"{max(0, min(255, round(channel))):02x}" for channel in rgb)


def interpolate_color(value, stops):
    value = max(0, min(1, value))
    for i in range(len(stops) - 1):
        left_value, left_color = stops[i]
        right_value, right_color = stops[i + 1]
        if left_value <= value <= right_value:
            local = (value - left_value) / (right_value - left_value)
            left_rgb = hex_to_rgb(left_color)
            right_rgb = hex_to_rgb(right_color)
            return rgb_to_hex(
                tuple(
                    left_rgb[channel]
                    + (right_rgb[channel] - left_rgb[channel]) * local
                    for channel in range(3)
                )
            )
    return stops[-1][1]


def finite_values(geojson, key):
    return [
        feature["properties"][key]
        for feature in geojson["features"]
        if isinstance(feature["properties"].get(key), (int, float))
        and math.isfinite(feature["properties"][key])
    ]


def add_colors(geojson):
    configs = {
        "part_mal_inscrits": [(0, "#f7fbff"), (0.5, "#6baed6"), (1, "#08306b")],
        "part_melenchon_exprimes": [(0, "#ffffe5"), (0.5, "#41ab5d"), (1, "#005a32")],
        "score_croise": [(0, "#b2182b"), (0.5, "#f7f7a6"), (1, "#1a9850")],
    }
    ranges = {
        key: (min(values), max(values))
        for key in configs
        for values in [finite_values(geojson, key)]
        if values
    }
    for feature in geojson["features"]:
        props = feature["properties"]
        colors = {}
        for key, stops in configs.items():
            if key not in props or key not in ranges:
                colors[key] = "#d9d9d9"
                continue
            low, high = ranges[key]
            value = 0.5 if high == low else (props[key] - low) / (high - low)
            colors[key] = interpolate_color(value, stops)
        props["_color_mal"] = colors["part_mal_inscrits"]
        props["_color_melenchon"] = colors["part_melenchon_exprimes"]
        props["_color_cross"] = colors["score_croise"]


def point_line_distance(point, start, end):
    if start == end:
        return math.dist(point, start)
    dx = end[0] - start[0]
    dy = end[1] - start[1]
    projection = (
        (point[0] - start[0]) * dx + (point[1] - start[1]) * dy
    ) / (dx * dx + dy * dy)
    projection = max(0, min(1, projection))
    nearest = (start[0] + projection * dx, start[1] + projection * dy)
    return math.dist(point, nearest)


def simplify_line(points, tolerance):
    if len(points) <= 2:
        return points
    start, end = points[0], points[-1]
    distances = [point_line_distance(point, start, end) for point in points[1:-1]]
    max_distance = max(distances, default=0)
    if max_distance <= tolerance:
        return [start, end]
    split = distances.index(max_distance) + 1
    return simplify_line(points[: split + 1], tolerance)[:-1] + simplify_line(
        points[split:], tolerance
    )


def simplify_ring(ring, tolerance):
    if len(ring) < 5:
        return ring
    open_ring = ring[:-1] if ring[0] == ring[-1] else ring
    simplified = simplify_line(open_ring, tolerance)
    if len(simplified) < 3:
        return ring
    return simplified + [simplified[0]]


def simplified_geojson(geojson, tolerance=0.01):
    result = json.loads(json.dumps(geojson))
    for feature in result["features"]:
        geometry = feature["geometry"]
        if geometry["type"] == "Polygon":
            geometry["coordinates"] = [
                simplify_ring(ring, tolerance) for ring in geometry["coordinates"]
            ]
        elif geometry["type"] == "MultiPolygon":
            geometry["coordinates"] = [
                [simplify_ring(ring, tolerance) for ring in polygon]
                for polygon in geometry["coordinates"]
            ]
    return result


def map_html(geojson, mode, title, subtitle, variable_label, source_note):
    feature_data = json.dumps(geojson, ensure_ascii=False, separators=(",", ":"))
    color_key = {
        "mal": "_color_mal",
        "melenchon": "_color_melenchon",
        "cross": "_color_cross",
    }[mode]
    legend = {
        "mal": "Faible part de mal-inscrits -> forte part",
        "melenchon": "Faible vote Mélenchon -> fort vote Mélenchon",
        "cross": "Rouge : faible sur au moins un axe · Vert : fort sur les deux axes",
    }[mode]
    return f"""<!doctype html>
<html lang="fr">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(title)}</title>
  <style>
    body {{ margin:0; font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif; color:#202124; background:#f8f7f3; }}
    header {{ padding:18px 22px 12px; background:#fff; border-bottom:1px solid #ddd9d0; }}
    h1 {{ margin:0; font-size:22px; font-weight:700; letter-spacing:0; }}
    header p {{ margin:6px 0 0; font-size:14px; color:#5f6368; }}
    main {{ display:grid; grid-template-columns:minmax(0,1fr) 320px; min-height:calc(100vh - 82px); }}
    #map {{ width:100%; height:calc(100vh - 82px); background:#e8eef1; }}
    aside {{ padding:18px; background:#fff; border-left:1px solid #ddd9d0; overflow:auto; }}
    .legend-bar {{ height:16px; border-radius:3px; margin:10px 0 6px; background:linear-gradient(90deg,#b2182b,#f7f7a6,#1a9850); }}
    .legend-bar.mal {{ background:linear-gradient(90deg,#f7fbff,#6baed6,#08306b); }}
    .legend-bar.melenchon {{ background:linear-gradient(90deg,#ffffe5,#41ab5d,#005a32); }}
    .metric {{ margin-top:16px; font-size:14px; line-height:1.45; }}
    .metric strong {{ display:block; font-size:13px; color:#5f6368; font-weight:600; }}
    .note {{ margin-top:20px; font-size:12px; line-height:1.45; color:#686868; }}
    svg {{ width:100%; height:100%; }}
    path.department {{ stroke:#fff; stroke-width:.7; cursor:pointer; transition:opacity 120ms ease,stroke-width 120ms ease; }}
    path.department:hover {{ opacity:.82; stroke:#1f1f1f; stroke-width:1.3; }}
    @media (max-width:820px) {{ main {{ grid-template-columns:1fr; }} #map {{ height:68vh; }} aside {{ border-left:0; border-top:1px solid #ddd9d0; }} }}
  </style>
</head>
<body>
  <header>
    <h1>{html.escape(title)}</h1>
    <p>{html.escape(subtitle)}</p>
  </header>
  <main>
    <section id="map" aria-label="Carte des départements"></section>
    <aside>
      <h2 id="dept-title">Survolez un département</h2>
      <div class="legend-bar {html.escape(mode)}"></div>
      <p>{html.escape(legend)}</p>
      <div class="metric"><strong>Variable cartographiée</strong>{html.escape(variable_label)}</div>
      <div id="details" class="metric"></div>
      <p class="note">{html.escape(source_note)}</p>
    </aside>
  </main>
  <script>
    const geojson = {feature_data};
    const colorKey = "{color_key}";
    const details = document.getElementById("details");
    const title = document.getElementById("dept-title");
    const map = document.getElementById("map");
    const width = 960;
    const height = 900;
    const svg = document.createElementNS("http://www.w3.org/2000/svg", "svg");
    svg.setAttribute("viewBox", "0 0 " + width + " " + height);
    svg.setAttribute("role", "img");
    map.appendChild(svg);

    const lonLat = [];
    for (const feature of geojson.features) {{
      walkCoordinates(feature.geometry.coordinates, pair => lonLat.push(pair));
    }}
    const lonMin = Math.min(...lonLat.map(d => d[0]));
    const lonMax = Math.max(...lonLat.map(d => d[0]));
    const latMin = Math.min(...lonLat.map(d => d[1]));
    const latMax = Math.max(...lonLat.map(d => d[1]));
    const margin = 24;

    function project(pair) {{
      const x = margin + (pair[0] - lonMin) / (lonMax - lonMin) * (width - margin * 2);
      const y = margin + (latMax - pair[1]) / (latMax - latMin) * (height - margin * 2);
      return [x, y];
    }}
    function walkCoordinates(coords, cb) {{
      if (typeof coords[0] === "number") cb(coords);
      else for (const child of coords) walkCoordinates(child, cb);
    }}
    function ringPath(ring) {{
      return ring.map((point, index) => {{
        const [x, y] = project(point);
        return (index === 0 ? "M" : "L") + x.toFixed(2) + "," + y.toFixed(2);
      }}).join(" ") + " Z";
    }}
    function geometryPath(geometry) {{
      if (geometry.type === "Polygon") return geometry.coordinates.map(ringPath).join(" ");
      if (geometry.type === "MultiPolygon") return geometry.coordinates.flatMap(poly => poly.map(ringPath)).join(" ");
      return "";
    }}
    function pct(value) {{
      return typeof value === "number" ? (value * 100).toFixed(1).replace(".", ",") + " %" : "n.d.";
    }}
    function integer(value) {{
      return typeof value === "number" ? Math.round(value).toLocaleString("fr-FR") : "n.d.";
    }}
    function show(feature) {{
      const p = feature.properties;
      title.textContent = p.libelle_departement + " (" + p.code_departement + ")";
      details.innerHTML = `
        <strong>Mal-inscription</strong>${{pct(p.part_mal_inscrits)}}
        <br><br><strong>Vote Mélenchon / exprimés</strong>${{pct(p.part_melenchon_exprimes)}}
        <br><br><strong>Vote Mélenchon / inscrits</strong>${{pct(p.part_melenchon_inscrits)}}
        <br><br><strong>Voix Mélenchon</strong>${{integer(p.voix_melenchon)}}
        <br><br><strong>Exprimés ministère</strong>${{integer(p.exprimes)}}
        <br><br><strong>Score croisé</strong>${{typeof p.score_croise === "number" ? p.score_croise.toFixed(3).replace(".", ",") : "n.d."}}
        <br><br><strong>Source vote</strong>${{p.source_melenchon || "n.d."}}
      `;
    }}
    for (const feature of geojson.features) {{
      const path = document.createElementNS("http://www.w3.org/2000/svg", "path");
      path.setAttribute("d", geometryPath(feature.geometry));
      path.setAttribute("class", "department");
      path.setAttribute("fill", feature.properties[colorKey] || "#d9d9d9");
      path.addEventListener("mouseenter", () => show(feature));
      path.addEventListener("focus", () => show(feature));
      path.setAttribute("tabindex", "0");
      path.setAttribute("aria-label", feature.properties.libelle_departement);
      svg.appendChild(path);
    }}
  </script>
</body>
</html>
"""


def write_maps(geojson):
    source_note = (
        "Mal-inscription : INSEE Première n°1986, Figure 4. "
        "Vote Mélenchon : ministère de l'Intérieur via Hexagonal, présidentielle 2022 tour 1. "
        "Le score croisé est la moyenne géométrique des deux rangs percentiles. "
        "Fond : GeoJSON simplifié des départements avec outre-mer rapprochée. "
        "Mayotte est affichée sans donnée de mal-inscription dans la source INSEE."
    )
    MAPS["mal"].write_text(
        map_html(
            geojson,
            "mal",
            "Mal-inscription en 2022 par département",
            "Part des inscrits dans une autre commune que leur commune de résidence principale.",
            "Part de mal-inscrits",
            source_note,
        ),
        encoding="utf-8",
    )
    MAPS["melenchon"].write_text(
        map_html(
            geojson,
            "melenchon",
            "Vote Mélenchon en 2022 par département",
            "Premier tour de l'élection présidentielle, part des suffrages exprimés.",
            "Part Mélenchon parmi les exprimés",
            source_note,
        ),
        encoding="utf-8",
    )
    MAPS["cross"].write_text(
        map_html(
            geojson,
            "cross",
            "Croisement mal-inscription x vote Mélenchon en 2022",
            "Rouge : faible sur au moins une dimension. Vert : élevé simultanément sur les deux dimensions.",
            "Moyenne géométrique des rangs de mal-inscription et de vote Mélenchon",
            source_note,
        ),
        encoding="utf-8",
    )
    MAPS["cross_web"].write_text(
        map_html(
            simplified_geojson(geojson),
            "cross",
            "Croisement mal-inscription x vote Mélenchon en 2022",
            "Rouge : faible sur au moins une dimension. Vert : élevé simultanément sur les deux dimensions.",
            "Moyenne géométrique des rangs de mal-inscription et de vote Mélenchon",
            source_note + " Version web : contours géographiques simplifiés, données inchangées.",
        ),
        encoding="utf-8",
    )


def main():
    ensure_dirs()
    mal_by_dep = extract_mal_inscription()
    melenchon_by_dep = aggregate_melenchon_t1()
    geojson = enrich_geojson(mal_by_dep, melenchon_by_dep)
    add_colors(geojson)
    json.dump(
        geojson,
        JOINED_GEOJSON.open("w", encoding="utf-8"),
        ensure_ascii=False,
        separators=(",", ":"),
    )
    write_maps(geojson)
    print(f"Mal-inscription CSV: {MAL_INSCRIPTION_CSV}")
    print(f"Vote Mélenchon CSV: {MELENCHON_CSV}")
    print(f"GeoJSON joint: {JOINED_GEOJSON}")
    for path in MAPS.values():
        print(f"Carte: {path}")


if __name__ == "__main__":
    main()
