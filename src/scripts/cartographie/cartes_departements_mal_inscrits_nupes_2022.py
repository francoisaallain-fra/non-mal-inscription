import csv
import html
import json
import math
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


ROOT = Path(__file__).resolve().parents[3]
INSEE_XLSX = ROOT / "data/01_raw/insee/donnees_insee_premiere_n1986.xlsx"
LEGIS_T1 = ROOT / "data/01_raw/ministere_interieur/2022-legislatives-1-bureau_de_vote.csv"
LEGIS_T2 = ROOT / "data/01_raw/ministere_interieur/2022-legislatives-2-bureau_de_vote.csv"
LEGIS_2022_NUANCES = ROOT / "data/01_raw/legis_2022/2022-legislatives-nuances.csv"
DEPARTEMENTS_GEOJSON = ROOT / "data/04_analysis/geodata/departements-2022.geojson"

MAL_INSCRIPTION_CSV = (
    ROOT / "data/04_analysis/insee/2022-mal-inscription-departements.csv"
)
NUPES_CSV = (
    ROOT / "data/04_analysis/elections/2022-legislatives-nupes-departements-t1.csv"
)
NUPES_T1_MINISTRY_CSV = (
    ROOT
    / "data/04_analysis/elections/2022-legislatives-nupes-departements-t1-ministere-legis2022.csv"
)
NUPES_T2_CSV = (
    ROOT / "data/04_analysis/elections/2022-legislatives-nupes-departements-t2.csv"
)
WIKIPEDIA_OVERRIDES_CSV = (
    ROOT
    / "data/04_analysis/elections/2022-legislatives-nupes-wikipedia-overrides.csv"
)
JOINED_GEOJSON = (
    ROOT / "data/04_analysis/cartographie/2022-departements-mal-inscription-nupes.geojson"
)
JOINED_T1_MINISTRY_GEOJSON = (
    ROOT
    / "data/04_analysis/cartographie/2022-departements-mal-inscription-nupes-t1-ministere-legis2022.geojson"
)
JOINED_T2_GEOJSON = (
    ROOT / "data/04_analysis/cartographie/2022-departements-mal-inscription-nupes-t2.geojson"
)
QA_UNMATCHED = ROOT / "data/04_analysis/qa/2022-departements-non-apparies.csv"
QA_UNMATCHED_T1_MINISTRY = (
    ROOT / "data/04_analysis/qa/2022-departements-non-apparies-t1-ministere-legis2022.csv"
)
QA_UNMATCHED_T2 = ROOT / "data/04_analysis/qa/2022-departements-non-apparies-t2.csv"
QA_TOTALS = ROOT / "data/04_analysis/qa/2022-controles-totaux-nupes-departements.csv"
QA_TOTALS_T1_MINISTRY = (
    ROOT
    / "data/04_analysis/qa/2022-controles-totaux-nupes-departements-t1-ministere-legis2022.csv"
)
QA_TOTALS_T2 = ROOT / "data/04_analysis/qa/2022-controles-totaux-nupes-departements-t2.csv"

MAPS = {
    "mal": ROOT / "maps/2022-mal-inscrits-departements.html",
    "nupes": ROOT / "maps/2022-vote-nupes-departements.html",
    "nupes_coalition_insee_style": ROOT
    / "maps/2022-vote-nupes-departements-coalition-style-insee.html",
    "cross_coalition": ROOT
    / "maps/2022-croisement-mal-inscrits-nupes-departements-coalition.html",
    "cross_coalition_web": ROOT
    / "maps/2022-croisement-mal-inscrits-nupes-departements-coalition-web.html",
    "cross_t1_ministry": ROOT
    / "maps/2022-croisement-mal-inscrits-nupes-departements-t1-ministere-legis2022.html",
    "cross_t2": ROOT
    / "maps/2022-croisement-mal-inscrits-nupes-departements-t2.html",
}

DEPARTEMENTS_OUTREMER = {
    "ZA": "971",
    "ZB": "972",
    "ZC": "973",
    "ZD": "974",
    "ZM": "976",
    "ZN": "988",
    "ZP": "987",
    "ZS": "975",
    "ZW": "986",
}

TERRITORY_INSET_BOXES = {
    "975": (-5.7, 43.55, -4.9, 44.15),
    "986": (-5.7, 42.75, -4.9, 43.35),
    "987": (-4.75, 43.55, -3.95, 44.15),
    "988": (-4.75, 42.75, -3.95, 43.35),
    "ZX": (-3.8, 43.55, -3.0, 44.15),
    "ZZ": (-3.8, 42.75, -3.0, 43.35),
}

OFFICIAL_OVERSEAS_NUPES_T1 = {
    "971": 25089,
    "972": 23286,
    "973": 6040,
    "974": 57373,
    "975": 782,
    "976": 2471,
    "986": 0,
    "987": 23490,
    "988": 0,
    "ZX": 0,
    "ZZ": 91576,
}

OFFICIAL_CORSICA_NUPES_T1 = {
    "2A": 4667,
    "2B": 4828,
}


def ensure_dirs():
    for path in [
        MAL_INSCRIPTION_CSV,
        NUPES_CSV,
        NUPES_T1_MINISTRY_CSV,
        NUPES_T2_CSV,
        JOINED_GEOJSON,
        JOINED_T1_MINISTRY_GEOJSON,
        JOINED_T2_GEOJSON,
        QA_UNMATCHED,
        QA_UNMATCHED_T1_MINISTRY,
        QA_UNMATCHED_T2,
        QA_TOTALS,
        QA_TOTALS_T1_MINISTRY,
        QA_TOTALS_T2,
        *MAPS.values(),
    ]:
        path.parent.mkdir(parents=True, exist_ok=True)


def code_departement(value):
    value = str(value).strip()
    return DEPARTEMENTS_OUTREMER.get(value, value.zfill(2))


def to_int(value):
    return int(value) if value else 0


def to_source_percent(value):
    value = str(value or "").strip().replace(",", ".")
    return float(value) / 100 if value else None


def normalize_circonscription(value):
    return str(value).strip().zfill(2)


def load_nupes_candidate_keys():
    keys = set()
    with LEGIS_2022_NUANCES.open("r", encoding="utf-8-sig", newline="") as stream:
        reader = csv.DictReader(stream)
        for row in reader:
            if not (
                row["CODAGE_NUPES_ENSEMBLE_OUTREMER"] == "NUPES"
                or row["CODAGE_LEGIS2022"].startswith("NUPES")
            ):
                continue
            keys.add(
                (
                    code_departement(row["DEP"]),
                    normalize_circonscription(row["NCIRCO"]),
                    to_int(row["panneau"]),
                )
            )
    return keys


def load_wikipedia_overrides():
    overrides = {}
    if not WIKIPEDIA_OVERRIDES_CSV.exists():
        return overrides

    with WIKIPEDIA_OVERRIDES_CSV.open("r", encoding="utf-8-sig", newline="") as stream:
        reader = csv.DictReader(stream)
        for row in reader:
            dep = row["code_departement"].strip()
            overrides[dep] = {
                "voix_nupes": to_int(row["voix_nupes_wikipedia_t1"]),
                "part_nupes_exprimes": to_source_percent(
                    row.get("pourcentage_nupes_wikipedia_t1")
                ),
                "source_nupes": row.get("source", "Wikipédia"),
                "methode_nupes": row.get(
                    "methode", "total_coalition_wikipedia"
                ),
            }
    return overrides


def extract_mal_inscription():
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl est requis pour extraire le fichier Insee brut. "
            "Les scripts qui réutilisent seulement les GeoJSON déjà produits "
            "peuvent importer ce module sans openpyxl."
        )
    wb = load_workbook(INSEE_XLSX, data_only=True)
    ws = wb["Figure 4"]

    rows = []
    for cells in ws.iter_rows(min_row=4, values_only=True):
        code, name, part_other_commune, part_same_commune = cells[:4]
        if not code or not name:
            continue
        rows.append(
            {
                "code_departement": code_departement(code),
                "libelle_departement": str(name).strip(),
                "part_mal_inscrits": f"{float(part_other_commune) / 100:.6f}",
                "part_inscrits_commune_residence": f"{float(part_same_commune) / 100:.6f}",
                "source": "INSEE Première n°1986, Figure 4",
            }
        )

    with MAL_INSCRIPTION_CSV.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)
    return {row["code_departement"]: row for row in rows}


CANDIDATE_HEADERS = {
    "N°Panneau",
    "Sexe",
    "Nom",
    "Prénom",
    "Nuance",
    "Voix",
    "% Voix/Ins",
    "% Voix/Exp",
}


def election_layout(headers):
    normalized = [header.rstrip(" 0123456789") for header in headers]
    repeated_names = []
    for header in normalized:
        if header in CANDIDATE_HEADERS and header not in repeated_names:
            repeated_names.append(header)
    first_candidate_idx = min(
        index for index, header in enumerate(normalized) if header in CANDIDATE_HEADERS
    )
    return {
        "departement_idx": normalized.index("Code du département"),
        "libelle_departement_idx": normalized.index("Libellé du département"),
        "circonscription_idx": normalized.index("Code de la circonscription"),
        "bureau_identity_indices": [
            normalized.index("Code de la circonscription"),
            normalized.index("Code de la commune"),
            normalized.index("Code du b.vote"),
        ],
        "inscrits_idx": normalized.index("Inscrits"),
        "votants_idx": normalized.index("Votants"),
        "blancs_idx": normalized.index("Blancs"),
        "nuls_idx": normalized.index("Nuls"),
        "exprimes_idx": normalized.index("Exprimés"),
        "first_candidate_idx": first_candidate_idx,
        "block_width": len(repeated_names),
        "panneau_offset": repeated_names.index("N°Panneau"),
        "nuance_offset": repeated_names.index("Nuance"),
        "voix_offset": repeated_names.index("Voix"),
    }


def candidate_blocks(line, layout):
    first = layout["first_candidate_idx"]
    width = layout["block_width"]
    for offset in range(first, len(line), width):
        if offset + width > len(line):
            continue
        yield {
            "panneau": line[offset + layout["panneau_offset"]].strip(),
            "nuance": line[offset + layout["nuance_offset"]].strip(),
            "voix": line[offset + layout["voix_offset"]].strip(),
        }


def aggregate_candidate_votes(source_path, turn, nupes_candidate_keys):
    totals = {}
    seen_bureaux_by_dep = {}
    with source_path.open("r", encoding="cp1252", newline="") as stream:
        reader = csv.reader(stream, delimiter=";")
        layout = election_layout(next(reader))
        for line in reader:
            dep = code_departement(line[layout["departement_idx"]])
            seen_bureaux = seen_bureaux_by_dep.setdefault(dep, set())
            bureau_key = tuple(
                line[index].strip()
                for index in layout["bureau_identity_indices"]
            )
            if bureau_key in seen_bureaux:
                continue
            seen_bureaux.add(bureau_key)

            bucket = totals.setdefault(
                dep,
                {
                    "code_departement": dep,
                    "libelle_departement": line[
                        layout["libelle_departement_idx"]
                    ],
                    "methode_nupes": (
                        "calcul_candidats" if turn == 1 else "calcul_candidats_t2"
                    ),
                    "source_nupes": (
                        f"Ministère de l'Intérieur T{turn} + Legis-2022"
                    ),
                    "source_exprimes": f"Ministère de l'Intérieur T{turn}",
                    "bureaux": 0,
                    "inscrits": 0,
                    "votants": 0,
                    "blancs": 0,
                    "nuls": 0,
                    "exprimes": 0,
                    "voix_nupes": 0,
                    "bureaux_avec_nupes": 0,
                },
            )
            bucket["bureaux"] += 1
            for field in ["inscrits", "votants", "blancs", "nuls", "exprimes"]:
                bucket[field] += to_int(line[layout[f"{field}_idx"]])

            has_nupes = False
            circonscription = normalize_circonscription(
                line[layout["circonscription_idx"]]
            )
            for candidate in candidate_blocks(line, layout):
                if not candidate["panneau"] or not candidate["voix"]:
                    continue
                candidate_key = (
                    dep,
                    circonscription,
                    to_int(candidate["panneau"]),
                )
                if (
                    candidate["nuance"] == "NUP"
                    or candidate_key in nupes_candidate_keys
                ):
                    bucket["voix_nupes"] += to_int(candidate["voix"])
                    has_nupes = True
            if has_nupes:
                bucket["bureaux_avec_nupes"] += 1
    return totals


def write_election_outputs(
    rows,
    output_csv,
    qa_totals,
    qa_scope,
    qa_method,
    calculate_national_rate=True,
):
    with output_csv.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)

    national = {
        "bureaux": sum(row["bureaux"] for row in rows),
        "inscrits": sum(row["inscrits"] for row in rows),
        "exprimes": sum(row["exprimes"] for row in rows),
        "voix_nupes": sum(row["voix_nupes"] for row in rows),
    }
    with qa_totals.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(
            stream,
            fieldnames=[
                "scope",
                "bureaux",
                "inscrits",
                "exprimes",
                "voix_nupes",
                "part_nupes_exprimes",
                "methode",
            ],
        )
        writer.writeheader()
        writer.writerow(
            {
                "scope": qa_scope,
                **national,
                "part_nupes_exprimes": (
                    f"{national['voix_nupes'] / national['exprimes']:.6f}"
                    if national["exprimes"] and calculate_national_rate
                    else ""
                ),
                "methode": qa_method,
            }
        )


def aggregate_nupes_t1(
    use_overrides=True,
    output_csv=NUPES_CSV,
    qa_totals=QA_TOTALS,
    qa_scope="France entière source ministère T1",
    qa_method="calcul candidats + totaux de coalition Wikipédia quand disponibles",
):
    totals = aggregate_candidate_votes(
        LEGIS_T1, 1, load_nupes_candidate_keys()
    )
    wikipedia_overrides = load_wikipedia_overrides() if use_overrides else {}
    rows = []
    for dep, row in sorted(totals.items()):
        if dep in wikipedia_overrides:
            override = wikipedia_overrides[dep]
            row["voix_nupes"] = override["voix_nupes"]
            row["part_nupes_exprimes_source"] = override["part_nupes_exprimes"]
            row["methode_nupes"] = override["methode_nupes"]
            row["source_nupes"] = override["source_nupes"]
        elif use_overrides and dep in OFFICIAL_OVERSEAS_NUPES_T1:
            row["voix_nupes"] = OFFICIAL_OVERSEAS_NUPES_T1[dep]
            row["part_nupes_exprimes_source"] = None
            row["methode_nupes"] = "total_coalition_officiel"
            row["source_nupes"] = "Totaux officiels de coalition"
        elif use_overrides and dep in OFFICIAL_CORSICA_NUPES_T1:
            row["voix_nupes"] = OFFICIAL_CORSICA_NUPES_T1[dep]
            row["part_nupes_exprimes_source"] = None
            row["methode_nupes"] = "total_coalition_officiel_corse"
            row["source_nupes"] = "Totaux officiels de coalition Corse"

        source_part = row.get("part_nupes_exprimes_source")
        if use_overrides and row["methode_nupes"] != "calcul_candidats":
            part_nupes_exprimes = source_part
            part_nupes_inscrits = None
            methode_part = (
                "pourcentage_source_coalition"
                if source_part is not None
                else "indisponible_sans_denominateur_coherent"
            )
        else:
            part_nupes_exprimes = (
                row["voix_nupes"] / row["exprimes"] if row["exprimes"] else None
            )
            part_nupes_inscrits = (
                row["voix_nupes"] / row["inscrits"] if row["inscrits"] else None
            )
            methode_part = "ratio_ministere"
        rows.append(
            {
                **row,
                "part_nupes_exprimes": (
                    f"{part_nupes_exprimes:.6f}"
                    if part_nupes_exprimes is not None
                    else ""
                ),
                "part_nupes_inscrits": (
                    f"{part_nupes_inscrits:.6f}"
                    if part_nupes_inscrits is not None
                    else ""
                ),
                "methode_part_nupes_exprimes": methode_part,
            }
        )

    if use_overrides:
        qa_method += (
            " ; taux national non calculé car les pourcentages de coalition "
            "proviennent de sources départementales distinctes"
        )
    write_election_outputs(
        rows,
        output_csv,
        qa_totals,
        qa_scope,
        qa_method,
        calculate_national_rate=not use_overrides,
    )
    return {row["code_departement"]: row for row in rows}


def aggregate_nupes_t2():
    totals = aggregate_candidate_votes(
        LEGIS_T2, 2, load_nupes_candidate_keys()
    )
    rows = []
    for _, row in sorted(totals.items()):
        rows.append(
            {
                **row,
                "part_nupes_exprimes": (
                    f"{row['voix_nupes'] / row['exprimes']:.6f}"
                    if row["exprimes"]
                    else ""
                ),
                "part_nupes_inscrits": (
                    f"{row['voix_nupes'] / row['inscrits']:.6f}"
                    if row["inscrits"]
                    else ""
                ),
                "methode_part_nupes_exprimes": "ratio_ministere",
            }
        )
    write_election_outputs(
        rows,
        NUPES_T2_CSV,
        QA_TOTALS_T2,
        "France entière source ministère T2",
        "calcul candidat par candidat sur les circonscriptions avec second tour",
    )
    return {row["code_departement"]: row for row in rows}


def percentile_scores(features, key):
    values = sorted(
        (feature["properties"].get(key), index)
        for index, feature in enumerate(features)
        if isinstance(feature["properties"].get(key), (int, float))
    )
    scores = {}
    if len(values) == 1:
        return {values[0][1]: 1.0}
    for rank, (_, index) in enumerate(values):
        scores[index] = rank / (len(values) - 1)
    return scores


def enrich_geojson(
    mal_by_dep,
    nupes_by_dep,
    output_geojson=JOINED_GEOJSON,
    qa_unmatched=QA_UNMATCHED,
):
    geojson = json.load(DEPARTEMENTS_GEOJSON.open(encoding="utf-8"))
    unmatched = []
    geo_codes = {feature["properties"].get("code") for feature in geojson["features"]}

    for feature in geojson["features"]:
        props = feature["properties"]
        dep = props.get("code")
        props["code_departement"] = dep
        props["libelle_departement"] = props.get("nom")

        mal = mal_by_dep.get(dep)
        nupes = nupes_by_dep.get(dep)
        if mal:
            props["part_mal_inscrits"] = float(mal["part_mal_inscrits"])
        else:
            unmatched.append({"code_departement": dep, "dataset": "mal_inscription"})

        if nupes:
            for key in [
                "bureaux",
                "inscrits",
                "exprimes",
                "voix_nupes",
                "bureaux_avec_nupes",
            ]:
                props[key] = int(nupes[key])
            props["methode_nupes"] = nupes["methode_nupes"]
            props["source_nupes"] = nupes["source_nupes"]
            props["source_exprimes"] = nupes["source_exprimes"]
            props["methode_part_nupes_exprimes"] = nupes[
                "methode_part_nupes_exprimes"
            ]
            if nupes["part_nupes_exprimes"]:
                props["part_nupes_exprimes"] = float(nupes["part_nupes_exprimes"])
            if nupes["part_nupes_inscrits"]:
                props["part_nupes_inscrits"] = float(nupes["part_nupes_inscrits"])
        else:
            unmatched.append({"code_departement": dep, "dataset": "vote_nupes"})

    for dep, nupes in sorted(nupes_by_dep.items()):
        if dep in geo_codes or dep not in TERRITORY_INSET_BOXES:
            continue
        x1, y1, x2, y2 = TERRITORY_INSET_BOXES[dep]
        props = {
            "code": dep,
            "nom": nupes["libelle_departement"],
            "code_departement": dep,
            "libelle_departement": nupes["libelle_departement"],
            "territoire_schematique": True,
            "absence_mal_inscription": "Non disponible dans l'étude INSEE n°1986",
        }
        for key in [
            "bureaux",
            "inscrits",
            "exprimes",
            "voix_nupes",
            "bureaux_avec_nupes",
        ]:
            props[key] = int(nupes[key])
        props["methode_nupes"] = nupes["methode_nupes"]
        props["source_nupes"] = nupes["source_nupes"]
        props["source_exprimes"] = nupes["source_exprimes"]
        props["methode_part_nupes_exprimes"] = nupes["methode_part_nupes_exprimes"]
        if nupes["part_nupes_exprimes"]:
            props["part_nupes_exprimes"] = float(nupes["part_nupes_exprimes"])
        if nupes["part_nupes_inscrits"]:
            props["part_nupes_inscrits"] = float(nupes["part_nupes_inscrits"])
        geojson["features"].append(
            {
                "type": "Feature",
                "properties": props,
                "geometry": {
                    "type": "Polygon",
                    "coordinates": [
                        [[x1, y1], [x2, y1], [x2, y2], [x1, y2], [x1, y1]]
                    ],
                },
            }
        )
        unmatched.append({"code_departement": dep, "dataset": "mal_inscription"})

    mal_scores = percentile_scores(geojson["features"], "part_mal_inscrits")
    nupes_scores = percentile_scores(geojson["features"], "part_nupes_exprimes")
    for index, feature in enumerate(geojson["features"]):
        props = feature["properties"]
        mal_score = mal_scores.get(index)
        nupes_score = nupes_scores.get(index)
        if mal_score is None or nupes_score is None:
            continue
        props["score_mal_inscription"] = round(mal_score, 6)
        props["score_nupes"] = round(nupes_score, 6)
        props["score_croise"] = round(math.sqrt(mal_score * nupes_score), 6)
        props["methode_score_croise"] = "moyenne_geometrique_des_rangs_percentiles"

    json.dump(
        geojson,
        output_geojson.open("w", encoding="utf-8"),
        ensure_ascii=False,
        separators=(",", ":"),
    )

    with qa_unmatched.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=["code_departement", "dataset"])
        writer.writeheader()
        writer.writerows(unmatched)

    return geojson


def hex_to_rgb(color):
    color = color.lstrip("#")
    return tuple(int(color[i : i + 2], 16) for i in (0, 2, 4))


def rgb_to_hex(rgb):
    return "#" + "".join(f"{max(0, min(255, round(channel))):02x}" for channel in rgb)


def interpolate_color(value, stops):
    value = max(0, min(1, value))
    for i in range(len(stops) - 1):
        left_value, left_color = stops[i]
        right_value, right_color = stops[i + 1]
        if left_value <= value <= right_value:
            local = (value - left_value) / (right_value - left_value)
            left_rgb = hex_to_rgb(left_color)
            right_rgb = hex_to_rgb(right_color)
            return rgb_to_hex(
                tuple(
                    left_rgb[channel]
                    + (right_rgb[channel] - left_rgb[channel]) * local
                    for channel in range(3)
                )
            )
    return stops[-1][1]


def finite_values(geojson, key):
    return [
        feature["properties"][key]
        for feature in geojson["features"]
        if isinstance(feature["properties"].get(key), (int, float))
        and math.isfinite(feature["properties"][key])
    ]


def quantile(values, q):
    values = sorted(values)
    if not values:
        return None
    position = (len(values) - 1) * q
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return values[int(position)]
    return values[lower] + (values[upper] - values[lower]) * (position - lower)


def add_nupes_insee_style_colors(geojson):
    values = finite_values(geojson, "part_nupes_exprimes")
    breaks = [quantile(values, q) for q in (0.25, 0.5, 0.75)]
    colors = ["#f8dbe0", "#f3aebd", "#b90f3e", "#68120d"]

    for feature in geojson["features"]:
        props = feature["properties"]
        value = props.get("part_nupes_exprimes")
        if not isinstance(value, (int, float)):
            props["_color_nupes_insee_style"] = "#d9d9d9"
            props["classe_nupes_insee_style"] = "n.d."
            continue
        if value < breaks[0]:
            index = 0
            label = f"Moins de {breaks[0] * 100:.1f} %"
        elif value < breaks[1]:
            index = 1
            label = f"De {breaks[0] * 100:.1f} à {breaks[1] * 100:.1f} %"
        elif value < breaks[2]:
            index = 2
            label = f"De {breaks[1] * 100:.1f} à {breaks[2] * 100:.1f} %"
        else:
            index = 3
            label = f"Plus de {breaks[2] * 100:.1f} %"
        props["_color_nupes_insee_style"] = colors[index]
        props["classe_nupes_insee_style"] = label.replace(".", ",")


def add_colors(geojson):
    configs = {
        "part_mal_inscrits": [(0, "#f7fbff"), (0.5, "#6baed6"), (1, "#08306b")],
        "part_nupes_exprimes": [(0, "#ffffe5"), (0.5, "#41ab5d"), (1, "#005a32")],
        "score_croise": [(0, "#b2182b"), (0.5, "#f7f7a6"), (1, "#1a9850")],
    }

    ranges = {
        key: (min(values), max(values))
        for key in configs
        for values in [finite_values(geojson, key)]
        if values
    }

    for feature in geojson["features"]:
        props = feature["properties"]
        colors = {}
        for key, stops in configs.items():
            if key not in props or key not in ranges:
                colors[key] = "#d9d9d9"
                continue
            low, high = ranges[key]
            value = 0.5 if high == low else (props[key] - low) / (high - low)
            colors[key] = interpolate_color(value, stops)
        props["_color_mal"] = colors["part_mal_inscrits"]
        props["_color_nupes"] = colors["part_nupes_exprimes"]
        props["_color_cross"] = colors["score_croise"]
    add_nupes_insee_style_colors(geojson)


def format_percent(value):
    if not isinstance(value, (int, float)):
        return "n.d."
    return f"{value * 100:.1f} %"


def point_line_distance(point, start, end):
    if start == end:
        return math.dist(point, start)
    dx = end[0] - start[0]
    dy = end[1] - start[1]
    projection = (
        (point[0] - start[0]) * dx + (point[1] - start[1]) * dy
    ) / (dx * dx + dy * dy)
    projection = max(0, min(1, projection))
    nearest = (start[0] + projection * dx, start[1] + projection * dy)
    return math.dist(point, nearest)


def simplify_line(points, tolerance):
    if len(points) <= 2:
        return points
    start, end = points[0], points[-1]
    distances = [
        point_line_distance(point, start, end) for point in points[1:-1]
    ]
    max_distance = max(distances, default=0)
    if max_distance <= tolerance:
        return [start, end]
    split = distances.index(max_distance) + 1
    return (
        simplify_line(points[: split + 1], tolerance)[:-1]
        + simplify_line(points[split:], tolerance)
    )


def simplify_ring(ring, tolerance):
    if len(ring) < 5:
        return ring
    open_ring = ring[:-1] if ring[0] == ring[-1] else ring
    simplified = simplify_line(open_ring, tolerance)
    if len(simplified) < 3:
        return ring
    return simplified + [simplified[0]]


def simplified_geojson(geojson, tolerance=0.01):
    result = json.loads(json.dumps(geojson))
    for feature in result["features"]:
        geometry = feature["geometry"]
        if geometry["type"] == "Polygon":
            geometry["coordinates"] = [
                simplify_ring(ring, tolerance)
                for ring in geometry["coordinates"]
            ]
        elif geometry["type"] == "MultiPolygon":
            geometry["coordinates"] = [
                [
                    simplify_ring(ring, tolerance)
                    for ring in polygon
                ]
                for polygon in geometry["coordinates"]
            ]
    return result


def nupes_insee_style_legend_html(geojson):
    colors = ["#f8dbe0", "#f3aebd", "#b90f3e", "#68120d"]
    labels = []
    for color in colors:
        label = next(
            (
                feature["properties"].get("classe_nupes_insee_style")
                for feature in geojson["features"]
                if feature["properties"].get("_color_nupes_insee_style") == color
            ),
            None,
        )
        if label:
            labels.append((color, label))
    if not labels:
        return ""
    items = "\n".join(
        f'<li><span class="swatch" style="background:{color}"></span>{html.escape(label)}</li>'
        for color, label in labels
    )
    return f'<ul class="class-legend">{items}</ul>'


def map_html(geojson, mode, title, subtitle, variable_label, source_note):
    feature_data = json.dumps(geojson, ensure_ascii=False, separators=(",", ":"))
    color_key = {
        "mal": "_color_mal",
        "nupes": "_color_nupes",
        "nupes_insee_style": "_color_nupes_insee_style",
        "cross": "_color_cross",
    }[mode]
    legend = {
        "mal": "Faible part de mal-inscrits → forte part",
        "nupes": "Faible vote NUPES → fort vote NUPES",
        "nupes_insee_style": "Rose clair : faible vote NUPES · Bordeaux : fort vote NUPES",
        "cross": "Rouge : faible sur au moins un axe · Vert : fort sur les deux axes",
    }[mode]
    class_legend = (
        nupes_insee_style_legend_html(geojson)
        if mode == "nupes_insee_style"
        else ""
    )
    class_legend_block = f"      {class_legend}\n" if class_legend else ""

    return f"""<!doctype html>
<html lang="fr">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(title)}</title>
  <style>
    body {{
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      color: #202124;
      background: #f8f7f3;
    }}
    header {{
      padding: 18px 22px 12px;
      background: #ffffff;
      border-bottom: 1px solid #ddd9d0;
    }}
    h1 {{
      margin: 0;
      font-size: 22px;
      font-weight: 700;
      letter-spacing: 0;
    }}
    header p {{
      margin: 6px 0 0;
      font-size: 14px;
      color: #5f6368;
    }}
    main {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) 320px;
      min-height: calc(100vh - 82px);
    }}
    #map {{
      width: 100%;
      height: calc(100vh - 82px);
      background: #e8eef1;
    }}
    aside {{
      padding: 18px;
      background: #ffffff;
      border-left: 1px solid #ddd9d0;
      overflow: auto;
    }}
    .legend-bar {{
      height: 16px;
      border-radius: 3px;
      margin: 10px 0 6px;
      background: linear-gradient(90deg, #b2182b, #f7f7a6, #1a9850);
    }}
    .legend-bar.mal {{
      background: linear-gradient(90deg, #f7fbff, #6baed6, #08306b);
    }}
    .legend-bar.nupes {{
      background: linear-gradient(90deg, #ffffe5, #41ab5d, #005a32);
    }}
    .legend-bar.nupes_insee_style {{
      background: linear-gradient(90deg, #f8dbe0 0 25%, #f3aebd 25% 50%, #b90f3e 50% 75%, #68120d 75% 100%);
    }}
    .class-legend {{
      list-style: none;
      padding: 0;
      margin: 12px 0 0;
      font-size: 14px;
      line-height: 1.5;
    }}
    .class-legend li {{
      display: flex;
      align-items: center;
      gap: 10px;
      margin: 8px 0;
    }}
    .swatch {{
      width: 22px;
      height: 22px;
      border: 1px solid rgba(0, 0, 0, 0.08);
      flex: 0 0 22px;
    }}
    .metric {{
      margin-top: 16px;
      font-size: 14px;
      line-height: 1.45;
    }}
    .metric strong {{
      display: block;
      font-size: 13px;
      color: #5f6368;
      font-weight: 600;
    }}
    .note {{
      margin-top: 20px;
      font-size: 12px;
      line-height: 1.45;
      color: #686868;
    }}
    svg {{
      width: 100%;
      height: 100%;
    }}
    path.department {{
      stroke: #ffffff;
      stroke-width: 0.7;
      cursor: pointer;
      transition: opacity 120ms ease, stroke-width 120ms ease;
    }}
    path.department:hover {{
      opacity: 0.82;
      stroke: #1f1f1f;
      stroke-width: 1.3;
    }}
    @media (max-width: 820px) {{
      main {{
        grid-template-columns: 1fr;
      }}
      #map {{
        height: 68vh;
      }}
      aside {{
        border-left: 0;
        border-top: 1px solid #ddd9d0;
      }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>{html.escape(title)}</h1>
    <p>{html.escape(subtitle)}</p>
  </header>
  <main>
    <section id="map" aria-label="Carte des départements"></section>
    <aside>
      <h2 id="dept-title">Survolez un département</h2>
      <div class="legend-bar {html.escape(mode)}"></div>
      <p>{html.escape(legend)}</p>
{class_legend_block}
      <div class="metric"><strong>Variable cartographiée</strong>{html.escape(variable_label)}</div>
      <div id="details" class="metric"></div>
      <p class="note">{html.escape(source_note)}</p>
    </aside>
  </main>
  <script>
    const geojson = {feature_data};
    const colorKey = "{color_key}";
    const details = document.getElementById("details");
    const title = document.getElementById("dept-title");
    const map = document.getElementById("map");
    const width = 960;
    const height = 900;
    const svg = document.createElementNS("http://www.w3.org/2000/svg", "svg");
    svg.setAttribute("viewBox", "0 0 " + width + " " + height);
    svg.setAttribute("role", "img");
    map.appendChild(svg);

    const lonLat = [];
    for (const feature of geojson.features) {{
      walkCoordinates(feature.geometry.coordinates, pair => lonLat.push(pair));
    }}
    const lonMin = Math.min(...lonLat.map(d => d[0]));
    const lonMax = Math.max(...lonLat.map(d => d[0]));
    const latMin = Math.min(...lonLat.map(d => d[1]));
    const latMax = Math.max(...lonLat.map(d => d[1]));
    const margin = 24;

    function project(pair) {{
      const x = margin + (pair[0] - lonMin) / (lonMax - lonMin) * (width - margin * 2);
      const y = margin + (latMax - pair[1]) / (latMax - latMin) * (height - margin * 2);
      return [x, y];
    }}

    function walkCoordinates(coords, cb) {{
      if (typeof coords[0] === "number") {{
        cb(coords);
      }} else {{
        for (const child of coords) walkCoordinates(child, cb);
      }}
    }}

    function ringPath(ring) {{
      return ring.map((point, index) => {{
        const [x, y] = project(point);
        return (index === 0 ? "M" : "L") + x.toFixed(2) + "," + y.toFixed(2);
      }}).join(" ") + " Z";
    }}

    function geometryPath(geometry) {{
      if (geometry.type === "Polygon") {{
        return geometry.coordinates.map(ringPath).join(" ");
      }}
      if (geometry.type === "MultiPolygon") {{
        return geometry.coordinates.flatMap(poly => poly.map(ringPath)).join(" ");
      }}
      return "";
    }}

    function pct(value) {{
      return typeof value === "number" ? (value * 100).toFixed(1).replace(".", ",") + " %" : "n.d.";
    }}

    function integer(value) {{
      return typeof value === "number" ? Math.round(value).toLocaleString("fr-FR") : "n.d.";
    }}

    function method(value) {{
      if (value === "total_coalition_wikipedia") return "Total Wikipédia de coalition";
      if (value === "total_coalition_wikipedia_t1") return "Total Wikipédia de coalition - premier tour";
      if (value === "total_coalition_wikipedia_t2") return "Total Wikipédia de coalition - second tour";
      if (value === "total_coalition_officiel") return "Total de coalition territorial";
      if (value === "total_coalition_officiel_corse") return "Total de coalition Corse";
      if (value === "calcul_candidats") return "Calcul candidat par candidat";
      if (value === "calcul_candidats_t2") return "Calcul candidat par candidat - second tour";
      return "n.d.";
    }}

    function show(feature) {{
      const p = feature.properties;
      title.textContent = p.libelle_departement + " (" + p.code_departement + ")";
      details.innerHTML = `
        <strong>Mal-inscription</strong>${{pct(p.part_mal_inscrits)}}
        <br><br><strong>Vote NUPES / exprimés</strong>${{pct(p.part_nupes_exprimes)}}
        <br><br><strong>Vote NUPES / inscrits</strong>${{pct(p.part_nupes_inscrits)}}
        <br><br><strong>Classe NUPES</strong>${{p.classe_nupes_insee_style || "n.d."}}
        <br><br><strong>Voix NUPES</strong>${{integer(p.voix_nupes)}}
        <br><br><strong>Exprimés ministère (contrôle)</strong>${{integer(p.exprimes)}}
        <br><br><strong>Méthode NUPES</strong>${{method(p.methode_nupes)}}
        <br><br><strong>Méthode du pourcentage</strong>${{p.methode_part_nupes_exprimes || "n.d."}}
        <br><br><strong>Source NUPES</strong>${{p.source_nupes || "n.d."}}
      `;
    }}

    for (const feature of geojson.features) {{
      const path = document.createElementNS("http://www.w3.org/2000/svg", "path");
      path.setAttribute("d", geometryPath(feature.geometry));
      path.setAttribute("class", "department");
      path.setAttribute("fill", feature.properties[colorKey] || "#d9d9d9");
      path.addEventListener("mouseenter", () => show(feature));
      path.addEventListener("focus", () => show(feature));
      path.setAttribute("tabindex", "0");
      path.setAttribute("aria-label", feature.properties.libelle_departement);
      svg.appendChild(path);
    }}
  </script>
</body>
</html>
"""


def write_maps(geojson, geojson_t1_ministry, geojson_t2):
    source_note = (
        "Mal-inscription : INSEE Première n°1986, Figure 4. "
        "Vote NUPES : ministère de l'Intérieur via Hexagonal, législatives 2022 tour 1 ; "
        "voix et pourcentages de coalition Wikipédia utilisés ensemble quand ils sont disponibles. "
        "Une correction sans pourcentage source cohérent reste sans taux cartographié. "
        "Le score croisé est la moyenne géométrique des deux rangs percentiles. "
        "Fond : GeoJSON simplifié des départements avec outre-mer rapprochée. "
        "Mayotte est affichée sans donnée de mal-inscription dans la source INSEE."
    )
    MAPS["mal"].write_text(
        map_html(
            geojson,
            "mal",
            "Mal-inscription en 2022 par département",
            "Part des inscrits dans une autre commune que leur commune de résidence principale.",
            "Part de mal-inscrits",
            source_note,
        ),
        encoding="utf-8",
    )
    MAPS["nupes"].write_text(
        map_html(
            geojson,
            "nupes",
            "Vote NUPES en 2022 par département",
            "Premier tour des élections législatives, part des suffrages exprimés.",
            "Part NUPES parmi les exprimés",
            source_note,
        ),
        encoding="utf-8",
    )
    MAPS["nupes_coalition_insee_style"].write_text(
        map_html(
            geojson,
            "nupes_insee_style",
            "Vote NUPES en 2022 par département - coalition contrôlée",
            "Premier tour : pourcentages Wikipédia de coalition quand disponibles, calcul ministère/Legis-2022 sinon.",
            "Part NUPES parmi les exprimés, en 4 classes par quartiles",
            source_note,
        ),
        encoding="utf-8",
    )
    MAPS["cross_coalition"].write_text(
        map_html(
            geojson,
            "cross",
            "Croisement mal-inscription x vote NUPES en 2022 - coalition contrôlée",
            "Rouge : faible sur au moins une dimension. Vert : élevé simultanément sur les deux dimensions.",
            "Moyenne géométrique des rangs de mal-inscription et de vote NUPES",
            source_note,
        ),
        encoding="utf-8",
    )
    MAPS["cross_coalition_web"].write_text(
        map_html(
            simplified_geojson(geojson),
            "cross",
            "Croisement mal-inscription x vote NUPES en 2022 - coalition contrôlée",
            "Rouge : faible sur au moins une dimension. Vert : élevé simultanément sur les deux dimensions.",
            "Moyenne géométrique des rangs de mal-inscription et de vote NUPES",
            source_note
            + " Version web : contours géographiques simplifiés, données inchangées.",
        ),
        encoding="utf-8",
    )
    MAPS["cross_t1_ministry"].write_text(
        map_html(
            geojson_t1_ministry,
            "cross",
            "Croisement mal-inscription x vote NUPES en 2022 - premier tour source ministère",
            "Vote NUPES au premier tour des législatives 2022, calculé avec les mêmes sources que la carte du second tour.",
            "Moyenne géométrique des rangs de mal-inscription et de vote NUPES au premier tour",
            "Mal-inscription : INSEE Première n°1986, Figure 4. Vote NUPES T1 : ministère de l'Intérieur via Hexagonal, législatives 2022 tour 1, avec identification des candidats NUPES par nuance NUP et Legis-2022. Cette carte n'utilise pas les totaux Wikipédia de coalition, afin d'être comparable avec la carte du second tour. Les territoires sans donnée INSEE de mal-inscription restent neutres.",
        ),
        encoding="utf-8",
    )
    MAPS["cross_t2"].write_text(
        map_html(
            geojson_t2,
            "cross",
            "Croisement mal-inscription x vote NUPES en 2022 - second tour",
            "Vote NUPES au second tour des législatives 2022, calculé sur les circonscriptions où un second tour a eu lieu.",
            "Moyenne géométrique des rangs de mal-inscription et de vote NUPES au second tour",
            "Mal-inscription : INSEE Première n°1986, Figure 4. Vote NUPES T2 : ministère de l'Intérieur via Hexagonal, législatives 2022 tour 2, avec identification des candidats NUPES par nuance NUP et Legis-2022. Les départements sans candidat NUPES au second tour peuvent apparaître à 0. Les territoires sans donnée INSEE de mal-inscription restent neutres.",
        ),
        encoding="utf-8",
    )


def main():
    ensure_dirs()
    mal_by_dep = extract_mal_inscription()
    nupes_by_dep = aggregate_nupes_t1()
    nupes_t1_ministry_by_dep = aggregate_nupes_t1(
        use_overrides=False,
        output_csv=NUPES_T1_MINISTRY_CSV,
        qa_totals=QA_TOTALS_T1_MINISTRY,
        qa_scope="France entière source ministère T1 sans overrides",
        qa_method="calcul candidat par candidat, mêmes sources que le second tour",
    )
    nupes_t2_by_dep = aggregate_nupes_t2()
    geojson = enrich_geojson(mal_by_dep, nupes_by_dep)
    geojson_t1_ministry = enrich_geojson(
        mal_by_dep,
        nupes_t1_ministry_by_dep,
        output_geojson=JOINED_T1_MINISTRY_GEOJSON,
        qa_unmatched=QA_UNMATCHED_T1_MINISTRY,
    )
    geojson_t2 = enrich_geojson(
        mal_by_dep,
        nupes_t2_by_dep,
        output_geojson=JOINED_T2_GEOJSON,
        qa_unmatched=QA_UNMATCHED_T2,
    )
    add_colors(geojson)
    add_colors(geojson_t1_ministry)
    add_colors(geojson_t2)
    json.dump(
        geojson,
        JOINED_GEOJSON.open("w", encoding="utf-8"),
        ensure_ascii=False,
        separators=(",", ":"),
    )
    json.dump(
        geojson_t1_ministry,
        JOINED_T1_MINISTRY_GEOJSON.open("w", encoding="utf-8"),
        ensure_ascii=False,
        separators=(",", ":"),
    )
    json.dump(
        geojson_t2,
        JOINED_T2_GEOJSON.open("w", encoding="utf-8"),
        ensure_ascii=False,
        separators=(",", ":"),
    )
    write_maps(geojson, geojson_t1_ministry, geojson_t2)
    print(f"Mal-inscription CSV: {MAL_INSCRIPTION_CSV}")
    print(f"Vote NUPES CSV: {NUPES_CSV}")
    print(f"Vote NUPES T1 source ministère CSV: {NUPES_T1_MINISTRY_CSV}")
    print(f"Vote NUPES T2 CSV: {NUPES_T2_CSV}")
    print(f"GeoJSON joint: {JOINED_GEOJSON}")
    print(f"GeoJSON joint T1 source ministère: {JOINED_T1_MINISTRY_GEOJSON}")
    print(f"GeoJSON joint T2: {JOINED_T2_GEOJSON}")
    for path in MAPS.values():
        print(f"Carte: {path}")


if __name__ == "__main__":
    main()
